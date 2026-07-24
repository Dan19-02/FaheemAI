#!/usr/bin/env python3
"""
Parse the GCC boards topic-level syllabus workbook into the RAG corpus module.

Input : scripts/data/GCC_Boards_Topic_Level_Syllabus_v2_VERIFIED.xlsx
Output: src/data/syllabusCorpus.ts  (a GENERATED TypeScript module)

Rules (mirror the workbook's own honesty tiers):
  - One corpus chunk per data row (Class | Subject | Unit/Topic | Sub-topics).
  - The Class column forward-fills (merged/blank continuation cells).
  - Grade ranges ("Class 9-10") are KEPT as ranges, one chunk each, so a
    chapter table of contents stays one strong retrieval unit instead of
    duplicate near-identical embeddings per grade.
  - Sheet-level Confidence gates inclusion: Verified/High ingest as-is,
    Medium ingests with an explicit unverified note appended, Low (SABIS)
    is EXCLUDED from the corpus (the board remains a profile option).
  - Board names emitted here are the canonical strings the frontend profile
    select uses; retrieval matches them case-insensitively but exactly.

Run:  python3 scripts/parse_syllabus.py    (from backend/, needs openpyxl)
"""
import json
import re
import sys
from datetime import date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip3 install openpyxl")

HERE = Path(__file__).resolve().parent
XLSX = HERE / "data" / "GCC_Boards_Topic_Level_Syllabus_v2_VERIFIED.xlsx"
OUT = HERE.parent / "src" / "data" / "syllabusCorpus.ts"

# sheet name -> (canonical board name, tier). Canonical names MUST match the
# frontend BOARDS options (Login.tsx / App.tsx) for the retrieval board boost.
SHEET_META = {
    "Cambridge CAIE": ("Cambridge (CAIE)", "international"),
    "Pearson Edexcel": ("Pearson Edexcel", "international"),
    "IB MYP + DP": ("IB", "international"),
    "American CCSS-NGSS-AP": ("American (US)", "international"),
    "CBSE NCERT": ("CBSE", "international"),
    "ICSE - ISC CISCE": ("ICSE / ISC", "international"),
    "French AEFE": ("French (AEFE)", "international"),
    "SABIS": ("SABIS", "international"),
    "Saudi Arabia MoE": ("Saudi Arabia MoE", "national"),
    "UAE MoE": ("UAE MoE", "national"),
    "Qatar MoEHE": ("Qatar MoEHE", "national"),
    "Kuwait MoE": ("Kuwait MoE", "national"),
    "Bahrain MoE": ("Bahrain MoE", "national"),
    "Oman MoE": ("Oman MoE", "national"),
}

MEDIUM_NOTE = " (Note: syllabus details from this source were not re-verified against the current year.)"


def clean(v):
    return re.sub(r"\s+", " ", str(v)).strip() if v is not None else ""


def normalize_grade(class_cell: str) -> str:
    """'Class 6' -> '6', 'Class 9-10' -> '9-10', 'Class 10 (Seconde)' -> '10'."""
    base = re.sub(r"\(.*?\)", "", class_cell)
    m = re.search(r"(\d{1,2})\s*-\s*(\d{1,2})", base)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    m = re.search(r"\d{1,2}", base)
    return m.group(0) if m else ""


def slug(s: str, maxlen: int = 96) -> str:
    s = re.sub(r"[^a-z0-9]+", "-", s.lower()).strip("-")
    return s[:maxlen].rstrip("-")


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    chunks = []
    ids = set()
    stats = []
    for sheet, (board, tier) in SHEET_META.items():
        ws = wb[sheet]
        header2 = clean(ws.cell(row=2, column=1).value)
        m = re.search(r"Confidence:\s*([A-Za-z]+)", header2)
        confidence = (m.group(1) if m else "Unknown").lower()
        if confidence == "low":
            stats.append((board, confidence, 0, "excluded from corpus"))
            continue

        n = 0
        cur_grade = ""
        for row in ws.iter_rows(min_row=5, values_only=True):
            c, subj, unit, detail = [clean(v) for v in (tuple(row) + (None,) * 4)[:4]]
            if c:
                cur_grade = normalize_grade(c)
            if not (subj or unit or detail) or not cur_grade:
                continue
            subject = subj[:80] or "General"
            topic = unit[:110] or "Overview"
            body = f"{topic}: {detail}" if detail else topic
            content = f"{board}, Grade {cur_grade}, {subject}. {body}"
            if confidence == "medium":
                content += MEDIUM_NOTE
            cid = "syl-" + slug(f"{board}-{cur_grade}-{subject}-{topic}")
            base_id, k = cid, 2
            while cid in ids:
                cid = f"{base_id}-{k}"
                k += 1
            ids.add(cid)
            chunks.append({
                "id": cid,
                "subject": subject,
                "topic": topic,
                "board": board,
                "grade": cur_grade,
                "content": content,
            })
            n += 1
        stats.append((board, confidence, n, tier))

    lines = ",\n".join("  " + json.dumps(c, ensure_ascii=False) for c in chunks)
    stat_lines = "\n".join(f" *   {b:22} {conf:9} {n:4} chunks  ({note})" for b, conf, n, note in stats)
    OUT.write_text(
        f"""/**
 * GENERATED FILE - do not edit by hand.
 *
 * GCC boards topic-level syllabus corpus (grades 6-12), one retrieval chunk
 * per verified workbook row. Regenerate after updating the workbook with:
 *   python3 scripts/parse_syllabus.py
 *
 * Source workbook: scripts/data/GCC_Boards_Topic_Level_Syllabus_v2_VERIFIED.xlsx
 * Generated {date.today().isoformat()} - {len(chunks)} chunks:
{stat_lines}
 *
 * Confidence gating: Verified/High as-is; Medium carries an explicit
 * unverified note; Low (SABIS) excluded (profile option only).
 */
export interface SyllabusChunk {{
  id: string;
  subject: string;
  topic: string;
  board: string;
  grade: string;
  content: string;
}}

export const SYLLABUS_CORPUS: SyllabusChunk[] = [
{lines},
];
""",
        encoding="utf-8",
    )
    print(f"Wrote {OUT.relative_to(HERE.parent)} with {len(chunks)} chunks")
    for b, conf, n, note in stats:
        print(f"  {b:22} {conf:9} {n:4}  {note}")


if __name__ == "__main__":
    main()
